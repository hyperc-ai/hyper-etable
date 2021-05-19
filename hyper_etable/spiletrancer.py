"""
Convert the HyperC objects back to Excel spreadsheet
"""

import hyperc.xtj
from collections import defaultdict
import schedula as sh
from formulas.cell import CellWrapper
import os.path
import openpyxl


def to_dict(self):
    nodes = {
        k: d['value']
        for k, d in self.dsp.default_values.items()
        if not isinstance(k, sh.Token)
    }
    nodes = {
        k: isinstance(v, str) and v.startswith('=') and '="%s"' % v or v
        for k, v in nodes.items() if v != [[sh.EMPTY]]
    }
    for d in self.dsp.function_nodes.values():
        fun = d['function']
        if isinstance(fun, CellWrapper):
            nodes.update(dict.fromkeys(d['outputs'], fun.__name__))
    return nodes


class SpileTrancer:
    def __init__(self, filename, xl_model, static_objects, table_records=None):
        self.static_objects = static_objects
        self.filename = filename
        self.xl_model = xl_model
        xl_model.finish()
        # self.xl_dict = xl_model.to_dict()
        self.xl_dict = to_dict(xl_model)
        self.wb = openpyxl.load_workbook(filename=filename, keep_vba=True)
        # self.wb = openpyxl.load_workbook(filename=filename)
    
    def gen_xl_addr(self, filename, sheetname, letter, rownum):
        filename = os.path.basename(filename)
        return f"'[{filename}]{sheetname}'!{letter}{rownum}".upper()

    def gen_opxl_addr(self, filename, sheetname, letter, rownum):
        filename = os.path.basename(filename)
        for shtn in self.wb.sheetnames:
            if shtn.upper() == sheetname.upper():
                sheetname = shtn
        return sheetname, f"{letter}{rownum}".upper()
    
    def calculate_excel(self):
        "Return full xlsx"
        # First, collect types of annotations

        attr_names_by_class = defaultdict(list)
        for key, val in sorted(type(self.static_objects).__annotations__.items()):
            if not hasattr(val, "__table_name__"):
                continue
            attr_names_by_class[val].append(key)
        
        all_inputs = {}  # this dict will rewrite all inputs
        
        for cls_, cells in attr_names_by_class.items():
            for cell in cells:
                rownum = int(cell.split("_")[-1])
                sheetname = cls_.__xl_sheet_name__
                for letter, type_ in cls_.__annotations__.items():
                    if len(letter) > 3:
                        continue  # ignore trash?
                    if not hasattr(getattr(self.static_objects, cell), letter):
                        continue
                    xl_cell_ref = self.gen_xl_addr(self.filename, sheetname, letter, rownum)
                    opxl_sht, opxl_cell_ref = self.gen_opxl_addr(self.filename, sheetname, letter, rownum)

                    # First, check what we currently have at this cell
                    if (type(self.xl_dict[xl_cell_ref]) == str and 
                            self.xl_dict[xl_cell_ref].upper().startswith("=TAKEIF")):
                        # TAKRIF -> replace default value
                        cellvalue = getattr(getattr(self.static_objects, cell), letter)
                        orig_cell = self.xl_dict[xl_cell_ref]
                        if type(cellvalue) == str: # if type_ == str:  # bug with type detector workaround
                            cellvalue = f'"{cellvalue}"'
                        # TODO: tokenize, rewrite, re-render
                        fm_cellvalue = f"=TAKEIF({cellvalue}, {orig_cell.split(',', 1)[1]}"
                        all_inputs[xl_cell_ref] = fm_cellvalue

                        # for opyxl
                        orig_opxl_cell = self.wb[opxl_sht][opxl_cell_ref]
                        orig_cell = orig_opxl_cell.value
                        opxl_cellvalue = f"=TAKEIF({cellvalue}, {orig_cell.split(',', 1)[1]}"
                        self.wb[opxl_sht][opxl_cell_ref].value = opxl_cellvalue
                    elif (type(self.xl_dict[xl_cell_ref]) == str and 
                            self.xl_dict[xl_cell_ref].upper().startswith("=SELECTFROMRANGE")):
                        cellvalue = getattr(getattr(self.static_objects, cell), letter)
                        if type(cellvalue) == str: # if type_ == str:  # bug with type detector workaround
                            cellvalue = f'"{cellvalue}"'
                        orig_opxl_cell = self.wb[opxl_sht][opxl_cell_ref]
                        orig_cell = orig_opxl_cell.value
                        if "," in orig_cell:
                            opxl_cellvalue = f"{orig_cell.split(',', 1)[0]}, {cellvalue})"
                        else:
                            opxl_cellvalue = f"{orig_cell.split(')', 1)[0]}, {cellvalue})"
                        self.wb[opxl_sht][opxl_cell_ref].value = opxl_cellvalue
                    elif (type(self.xl_dict[xl_cell_ref]) == str and 
                            self.xl_dict[xl_cell_ref].upper().startswith("=")):
                        pass
                    else:  # raw value? just write what we have
                        cellvalue = getattr(getattr(self.static_objects, cell), letter)
                        all_inputs[xl_cell_ref] = cellvalue
                        self.wb[opxl_sht][opxl_cell_ref] = cellvalue


        self.xl_model.calculate(inputs=all_inputs)

        # TODO: assert and double-check calculations
    
    def write(self, dir):
        self.wb.save(os.path.join(dir, os.path.basename(self.filename)))
        # self.xl_model.write(dirpath=dir)





