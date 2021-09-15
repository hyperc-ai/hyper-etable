from collections import defaultdict
from os import mkdir
import string
import ast
import formulas.excel
import formulas
import hyperc
import hyperc.util
import hyperc.settings
import hyper_etable.etable_transpiler
import hyper_etable.spiletrancer
import hyper_etable.cell_resolver
import hyper_etable.connector
import hyperc.xtj
import itertools
import sys
import time
import types
import collections
import copy
import os.path
import pathlib
import openpyxl
import hyper_etable.util
import hyper_etable.pysourcebuilder
import pydoc
import hyper_etable.meta_table

hyperc.settings.IGNORE_MISSING_ATTR_BRANCH = 1

def stack_code_gen_all(objects):
    l_all_hasattr_drop = []
    for cname, rows in objects.items():
        for idx, rowobj in rows.items():
            colnames = type(rowobj).__annotations__.keys()
            for col in colnames:
                if col.startswith("_"): continue
                if len(col) > 4: continue
                # if not "not_hasattr" in col: continue
                if not hasattr(rowobj, col): continue
                if getattr(rowobj, f"{col}_not_hasattr") == False: continue
                l_all_hasattr_drop.append(f"DATA.{cname}_{idx}.{col}_not_hasattr = True")

    drop_content = "\n    ".join(l_all_hasattr_drop)
    warrants = '\n    '.join(hyper_etable.etable_transpiler.generate_ne_warrants(drop_content))
    scode = f"""def _stack_drop():
    pass
    {warrants}
    {drop_content}"""
    return scode


def operator_name_to_operator(op):
    expand = {">": "greaterThan", ">=": "greaterThanOrEqual", "<": "lessThan", "<=": "lessThanOrEqual",
              "==": "equal", "!=": "notEqual"}
    for k, v in expand.items():
        if v == op:
            return k
    return None

# H 75-150 is green
def rgb_to_hsv(r, g, b):
    r, g, b = r/255.0, g/255.0, b/255.0
    mx = max(r, g, b)
    mn = min(r, g, b)
    df = mx-mn
    if mx == mn:
        h = 0
    elif mx == r:
        h = (60 * ((g-b)/df) + 360) % 360
    elif mx == g:
        h = (60 * ((b-r)/df) + 120) % 360
    elif mx == b:
        h = (60 * ((r-g)/df) + 240) % 360
    if mx == 0:
        s = 0
    else:
        s = (df/mx)*100
    v = mx*100
    return h, s, v


# Define TAKEIF formula
FUNCTIONS = formulas.get_functions()
def STUB_TAKEIF(default, *args):
    return default
FUNCTIONS["TAKEIF"] = STUB_TAKEIF

# Define SELECTFROMRANGE formula
def STUB_SELECTFROMRANGE(*args):
    return 0
FUNCTIONS["SELECTFROMRANGE"] = STUB_SELECTFROMRANGE

# Define WATCH formula
def STUB_WATCHTAKEIF(*args):
    return 0
FUNCTIONS["WATCHTAKEIF"] = STUB_WATCHTAKEIF

FUNCTIONS["side_effect"] = hyperc.side_effect

class EventNameHolder:
    ename: str
    def __init__(self) -> None:
        self.ename = ""
    def __str__(self):
        v = str(self.ename)
        if v.startswith('"') and v.endswith('"'):
            v = v[1:-1]
        return v

DATA = None
HCT_OBJECTS = None

class ETable:
    def __init__(self, filenames=None, project_name="my_project", has_header=True) -> None:
        self.has_header = has_header
        if filenames is not None:
            if isinstance(filenames,list):
                filenames = [pathlib.Path(f) for f in filenames]
                self.filename = filenames[0] #TODO currently only one file support
            else:
                self.filename = pathlib.Path(filenames)
            if 'xlsx' == os.path.splitext(self.filename)[1][1:].lower():
                self.enable_precalculation = False
            else:
                self.enable_precalculation = True
        else: 
            self.filename = filenames
            self.enable_precalculation = False

        self.STATIC_STORAGE_NAME = 'DATA'
        self.plan_data_prefix = 'DATA.'
        self.out_filename = ""
        APPENDIX = hyperc.settings.APPENDIX
        hyperc.settings.APPENDIX = hyperc.xtj.str_to_py(str(self.filename)) + "_" + project_name
        self.tempdir = hyperc.util.get_work_dir()
        hyperc.settings.APPENDIX = APPENDIX
        self.session_name = "etable_mod"
        self.mod = types.ModuleType(self.session_name)
        globals()[self.session_name] = self.mod
        sys.modules[self.session_name] = self.mod
        self.classes = {}
        self.objects = {}
        self.mod.side_effect = hyperc.side_effect
        self.mod.ensure_ne = hyperc.ensure_ne
        self.methods_classes = {}

        self.mod.DefinedTables = type("DefinedTables", (object, ), {})
        self.mod.DefinedTables.__annotations__ = {}
        self.mod.DefinedTables.__qualname__ = f"{self.session_name}.DefinedTables"
        self.mod.DEFINED_TABLES = self.mod.DefinedTables()
        self.methods_classes["DefinedTables"] = self.mod.DefinedTables

        self.mod.StaticObject = type("StaticObject", (object, ), {})
        self.mod.StaticObject.__annotations__ = {}
        self.mod.StaticObject.__qualname__ = f"{self.session_name}.StaticObject"
        self.mod.DATA = self.mod.StaticObject()
        self.mod.DATA.GOAL = False
        globals()[self.STATIC_STORAGE_NAME] = self.mod.DATA
        self.mod.HCT_OBJECTS = {}
        globals()['HCT_OBJECTS'] = self.mod.HCT_OBJECTS
        globals()['side_effect'] = hyperc.side_effect
        self.methods_classes["StaticObject"] = self.mod.StaticObject

        self.plan_log = []
        self.cells_value = {}
        self.range_resolver = None # will be initialized later in self.calulate() not in self.open_dump
        self.plan_or_invariants = None
        self.source_code = defaultdict(list)
        self.metadata = {}

    def get_cellvalue_by_cellname(self, cellname):
        filename, sheet, row, column = hyper_etable.etable_transpiler.split_cell(cellname) 
        py_table_name = hyperc.xtj.str_to_py(f'[{filename}]{sheet}')
        attrname = f"{py_table_name}_{row}"
        return getattr(getattr(self.mod.DATA, attrname), column.upper())

    def get_row_by_cellname(self, cellname):
        filename, sheet, row, column = hyper_etable.etable_transpiler.split_cell(cellname) 
        py_table_name = hyperc.xtj.str_to_py(f'[{filename}]{sheet}')
        attrname = f"{py_table_name}_{row}"
        return getattr(self.mod.DATA, attrname)

    def solver_call_simple(self,goal, extra_instantiations):
        return hyperc.solve(goal, globals_=self.methods_classes, extra_instantiations=extra_instantiations, work_dir=self.tempdir, 
                            addition_modules=[self.mod], metadata=self.metadata)

    def solver_call_simple_wo_exec(self):
        def gg(s, g, e):
            hyperc.solve(g, globals_=s.methods_classes, extra_instantiations=e, work_dir=s.tempdir, 
                            addition_modules=[s.mod], metadata=s.metadata)
        self.metadata = {"plan_steps": [], "plan_exec": []}
        gg(self,self.methods_classes[self.main_goal.name],
                                              list(filter(lambda x: isinstance(x, type), self.methods_classes.values())))
    def solver_call_plan_n_exec(self):
        def gg(s, g, e):
            hyperc.solve(g, globals_=s.methods_classes, extra_instantiations=e, work_dir=s.tempdir, 
                            addition_modules=[s.mod], metadata=s.metadata)
        self.metadata = {"plan_steps": [], "plan_exec": [], "force_exec": True}
        gg(self,self.methods_classes[self.main_goal.name],
                                              list(filter(lambda x: isinstance(x, type), self.methods_classes.values())))

    def solver_call_simple_with_exec(self):
        def gg(s, g, e):
            hyperc.solve(g, globals_=s.methods_classes, extra_instantiations=e, work_dir=s.tempdir, 
                            addition_modules=[s.mod], metadata=s.metadata)
        self.metadata = None
        gg(self,self.methods_classes[self.main_goal.name],
                                              list(filter(lambda x: isinstance(x, type), self.methods_classes.values())))  

    def solver_call(self,goal, extra_instantiations):
        self.metadata = {"plan_steps": [], "plan_exec": []}
        mod=self.mod
        DATA = mod.DATA
        globals_ = self.methods_classes
        ret = hyperc.solve(goal, globals_=globals_, extra_instantiations=extra_instantiations, work_dir=self.tempdir, 
                            addition_modules=[mod], metadata=self.metadata)
        step_counter = 1
        sheetmap = {s.upper():s for s in self.wb_values_only.sheetnames}
        ename = EventNameHolder()
        for step in self.metadata["plan_exec"]:
            substep_counter = 0
            orig_vars = {}
            for cellvar in step[0].orig_funcobject.effect_vars:
                if cellvar.is_range:
                    recid_ret = range(cellvar.number[0], cellvar.number[1]+1)
                    letter = cellvar.letter[0]
                else:
                    recid_ret = [cellvar.number]
                    letter = cellvar.letter
                for number in recid_ret:
                    cell = hyper_etable.cell_resolver.PlainCell(
                        filename=cellvar.cell.filename, sheet=cellvar.cell.sheet, letter=letter, number=number)
                    orig_vars[str(cell)] = self.get_cellvalue_by_cellname(str(cell))
            step[0](**step[1])
            for cellvar in step[0].orig_funcobject.effect_vars:
                if cellvar.is_range:
                    recid_ret = range(cellvar.number[0], cellvar.number[1]+1)
                    letter = cellvar.letter[0]
                else:
                    recid_ret = [cellvar.number]
                    letter = cellvar.letter
                for number in recid_ret:
                    cell = hyper_etable.cell_resolver.PlainCell(
                        filename=cellvar.cell.filename, sheet=cellvar.cell.sheet, letter=letter, number=number)
                    new_value = self.get_cellvalue_by_cellname(str(cell))
                    ftype = step[0].orig_funcobject.formula_type
                    if ftype == "TAKEIF":
                        explain_type = "Decide whether to update value"
                    elif ftype == "SELECTFROMRANGE":
                        explain_type = "Choose one value from a list"
                    else:
                        explain_type = "(formula recalculation)"

                    if new_value == orig_vars[str(cell)] and ftype == "TAKEIF":
                        continue
                    filename, sheet, row, column = hyper_etable.etable_transpiler.split_cell(str(cell))
                    leftmost = ""
                    for ltr in string.ascii_uppercase:
                        colval = self.wb_values_only[sheetmap[sheet]][f"{ltr}{row}"].value
                        if type(colval) == str and len(colval) > 0:
                            leftmost = colval
                            break
                    topmost = ""
                    i = 0
                    prev_empty = False
                    for r in self.wb_values_only[sheetmap[sheet]].rows:
                        colval = r[string.ascii_uppercase.index(column.upper())].value
                        if type(colval) == str and len(colval) > 0:
                            if prev_empty == True:
                                topmost = colval
                            prev_empty = False
                        elif colval == None or colval == "":
                            prev_empty = True
                        else:
                            prev_empty = False
                        i += 1
                        if i >= row: break
                    log_entry = [step_counter, explain_type, ename, leftmost, topmost,
                                f"'{sheet.upper()}'!{column.upper()}", row, orig_vars[str(cell)],
                                new_value,
                                "'"+",".join(step[0].orig_funcobject.formula_str).replace(f"[{filename.upper()}]", "")]
                    if ftype == "TAKEIF":
                        #TODO load string from openpyxl
                        # ename_l = list(step[0].orig_funcobject.sync_cell)
                        # if ename_l:
                        #     ename_str = ename_l[0]
                        #     if isinstance(ename_str, hyper_etable.etable_transpiler.StringLikeVariable):
                        #         ename_str = ename_str.cell_str
                        #     ename_str = str(ename_str)
                        #     if not ename_str.endswith('"') and not ename_str.startswith('"') and "!" in ename_str:
                        #         ename_str = self.get_cellvalue_by_cellname(ename_str)
                        #     ename.ename = ename_str
                        ename = EventNameHolder()
                    self.plan_log.append(log_entry)
                    substep_counter += 1
            step_counter += 1



    def dump_functions(self, code, filename):
        s_code = ''
        fn = f"{self.tempdir}/{filename}"
        with open(fn, "w+") as f:
            for func in code.values():
                s_code += str(func)
            f.write(s_code)
        f_code = compile(s_code, fn, 'exec')

        exec(f_code, self.mod.__dict__)
        for func_code in code.values():
            self.methods_classes[func_code.name] = self.mod.__dict__[func_code.name]
            self.methods_classes[func_code.name].orig_source = str(func_code)
            self.methods_classes[func_code.name].orig_funcobject = func_code


    def get_new_table(self, table_name, sheet):
        ThisTable = hyper_etable.meta_table.TableElementMeta(f'{table_name}_Class', (object,), {'__table_name__': table_name, '__xl_sheet_name__': sheet})
        ThisTable.__annotations__ = {'__table_name__': str}
        ThisTable.__touched_annotations__ = set()
        ThisTable.__annotations_type_set__ = defaultdict(set)
        self.mod.__dict__[f'{table_name}_Class'] = ThisTable
        self.classes[table_name] = ThisTable
        self.classes[table_name].__qualname__ = f"{self.session_name}.{table_name}_Class"
        self.mod.HCT_OBJECTS[table_name] = []
        return ThisTable

    def get_object_from_var(self, var):
        py_table_name = hyperc.xtj.str_to_py(f'[{var.filename}]{var.sheet}')
        return self.objects[py_table_name][var.number]

    # def add_row(row):

    def open_from(self, path, has_header=None, addition_python_files=[], external_classes_filename=None, proto='xlsx'):
        if has_header is not None:
            self.has_header = has_header
        goal_code_source = {}

        self.main_goal = hyper_etable.etable_transpiler.FunctionCode(name=f'hct_main_goal', is_goal=True)
        self.main_goal.operators[self.main_goal.name].append('assert DATA.GOAL == True')
        self.main_goal.operators[self.main_goal.name].append('pass')
        goal_code_source['main_goal'] = self.main_goal
        conn = None
        if proto.lower() == 'msapi':
            conn = hyper_etable.connector.MSAPIConnector(path, self.mod, has_header=has_header)
        elif proto.lower() == 'gsheet':
            conn = hyper_etable.connector.GSheetConnector(path, self.mod, has_header=has_header)
        elif proto.lower() == 'xlsx':
            conn = hyper_etable.connector.XLSXConnector(path, self.mod, has_header=has_header)
        if conn is None:
            raise ValueError(f'{proto} is not support')
        self.objects.update(conn.objects)

        self.load_external_classes(external_classes_filename)
        for py_table in self.mod.HCT_OBJECTS.values():
            for row in py_table:
                for attr, value in row.__default_init__.items():
                    if value == 'set()':
                       setattr(row, attr, set())
                    elif isinstance(value, str):
                        setattr(row, attr, value.strip('"'))
                    else:
                        setattr(row, attr, value)

        for clsv in self.classes.values():
            var_global_addidx_name = f'DATA.{clsv.__table_name__}_addidx'
            setattr(self.mod.DATA, f'{clsv.__table_name__}_addidx', 0)
            self.mod.StaticObject.__annotations__[f'{clsv.__table_name__}_addidx'] = int
            init_f_code = []
            init_pars = []
            if hyperc.settings.DEBUG:
                print(f" {clsv} -  {clsv.__annotations__}")
            # init_f_code.append(f"global DATA")
            init_f_code.append(f"self.addidx = {var_global_addidx_name}")
            init_f_code.append(f"{var_global_addidx_name} += 1")
            for par_name, par_type in clsv.__annotations__.items():
                if par_name in ['__table_name__', 'addidx']:
                    continue
                # Skip None type cell
                if par_type is None:
                    par_type = str
                    clsv.__annotations__[par_name] = str
                # init_f_code.append(
                    # f'self.{par_name}_not_hasattr = True')  # TODO: statically set to true instead of asking in parameters
                if par_name in clsv.__default_init__:
                    if clsv.__default_init__[par_name] == 'set()':
                        init_f_code.append(f'self.{par_name} = set() # set init "{par_name.upper()}" of table "{clsv.__table_name__}"')
                    else:
                        init_pars.append(f"hct_p_{par_name}:{par_type.__name__}={clsv.__default_init__[par_name]}")
                        init_f_code.append(f'self.{par_name} = hct_p_{par_name} # cell "{par_name.upper()}" of table "{clsv.__table_name__}"')
                else:
                    if not par_type in hyperc.xtj.DEFAULT_VALS:
                        raise TypeError(f"Could not resolve type for {clsv.__name__}.{par_name} (forgot to init cell?)")
                    init_pars.append(f"hct_p_{par_name}:{par_type.__name__}={hyperc.xtj.DEFAULT_VALS[par_type]}")
                    init_f_code.append(f'self.{par_name} = hct_p_{par_name} # cell "{par_name.upper()}" of table "{clsv.__table_name__}"')

            if len(init_f_code) == 0:
                continue
            init_f_code.append(f'side_effect(lambda: HCT_OBJECTS["{clsv.__table_name__}"].append(self))')
            init_f_code.append(f'side_effect(lambda: setattr(self, "__recid__", self.__class__.__recid_max__ + self.addidx))')
            init_f_code.append(f'side_effect(lambda: setattr(self, "__header_back_map__",  self.__class__.__header_back_map__))')
            init_f_code.append(f'side_effect(lambda: setattr(self, "__touched_annotations__",  set()))')
            init_f_code.append(f'side_effect(lambda: [self.__touched_annotations__.add(o) for o in self.__annotations__ if (not (o.startswith("__") and o.endswith("__")) and (o not in getattr(self.__class__,"__user_defined_annotations__", [])) and o != "addidx")])')
            c=f'side_effect(lambda: setattr(DATA, f"{clsv.__table_name__}_'
            init_f_code.append(c+'{self.__recid__}", self))')
            full_f_code = '\n    '.join(init_f_code)
            full_f_pars = ",".join(init_pars)
            full_code = f"def hct_f_init(self, {full_f_pars}):\n    {full_f_code}"
            
            fn = f"{self.tempdir}/hpy_init_{clsv.__name__}.py"
            open(fn, "w+").write(full_code)
            f_code = compile(full_code, fn, 'exec')
            exec(f_code, globals())
            clsv.__init__ = globals()["hct_f_init"]
            clsv.__init__.__name__ = "__init__"

        # Now generate init for static object
        self.mod.DATA.GOAL = False
        self.mod.StaticObject.__annotations__['GOAL'] = bool
        init_f_code = []
        for attr_name, attr_type in self.mod.StaticObject.__annotations__.items():
            init_f_code.append(f"self.{attr_name} = DATA.{attr_name}")  # if it does not ignore, fix it!
        self.mod.StaticObject.__annotations__['GOAL'] = bool
        if init_f_code:

            full_f_code = '\n    '.join(init_f_code)
            full_code = f"def hct_stf_init(self):\n    {full_f_code}"
            fn = f"{self.tempdir}/hpy_stf_init_{self.mod.StaticObject.__name__}.py"
            open(fn, "w+").write(full_code)
            f_code = compile(full_code, fn, 'exec')
            exec(f_code, self.mod.__dict__)
            self.mod.StaticObject.__init__ = self.mod.__dict__["hct_stf_init"]
            self.mod.StaticObject.__init__.__name__ = "__init__"

        # dump goals and actions
        self.dump_functions(goal_code_source, 'hpy_goals.py')

        # addition python code
        for code_file in addition_python_files:
            addition_code = open(code_file, "r").read()
            if addition_code.startswith("from"):  # workaround for module imports
                addition_code = "#"+addition_code
            f_code = compile(addition_code, code_file, 'exec')
            exec(f_code, self.mod.__dict__)
            for f_name in f_code.co_names:
                if "." in f_name: continue  # workaround for module names
                t = self.mod.__dict__.get(f_name, None)
                if isinstance(t, types.FunctionType) or isinstance(t, types.MethodType) or isinstance(t, type):
                    self.methods_classes[f_name] = self.mod.__dict__[f_name]
        # for f in self.mod.__dict__:
        #     if isinstance(self.mod.__dict__[f], types.FunctionType):
        #         self.methods_classes[f] = self.mod.__dict__[f]

        self.methods_classes.update(self.classes)


    def open_dump(self, has_header=None, addition_python_files=[], external_classes_filename=None):
        if has_header is not None:
            self.has_header = has_header
        xl_mdl = formulas.excel.ExcelModel()
        xl_mdl.loads(str(self.filename))
        self.stl = hyper_etable.spiletrancer.SpileTrancer(self.filename, xl_mdl, self.mod.DATA, plan_log=self.plan_log)
        filename_case_remap_workaround = {}

        goal_code_source = {}

        self.main_goal = hyper_etable.etable_transpiler.FunctionCode(name=f'hct_main_goal', is_goal=True)
        self.main_goal.operators[self.main_goal.name].append('assert DATA.GOAL == True')
        self.main_goal.operators[self.main_goal.name].append('pass')
        goal_code_source['main_goal'] = self.main_goal
        # Load used cell
        for wb_sheet in self.wb_values_only:
            sheet = wb_sheet.title
            filename = self.filename
            filename = filename_case_remap_workaround.get(filename, filename)
            # py_table_name = hyperc.xtj.str_to_py(f'[{filename}]{sheet}')
            py_table_name = hyperc.xtj.str_to_py(f'{sheet}') # warning only sheet in 
            header_map = {}
            header_back_map = {}
            if self.has_header:
                is_header = True
            else:
                is_header = False
            ThisTable = hyper_etable.meta_table.TableElementMeta(f'{py_table_name}_Class', (object,), {'__table_name__': py_table_name, '__xl_sheet_name__': sheet})
            ThisTable.__annotations__ = {'__table_name__': str, 'addidx': int}
            ThisTable.__header_back_map__ = header_back_map
            ThisTable.__user_defined_annotations__ = []
            ThisTable.__default_init__ = {}
            ThisTable.__touched_annotations__ = set()
            ThisTable.__annotations_type_set__ = defaultdict(set)
            self.mod.__dict__[f'{py_table_name}_Class'] = ThisTable
            self.classes[py_table_name] = ThisTable
            self.classes[py_table_name].__qualname__ = f"{self.session_name}.{py_table_name}_Class"
            self.mod.HCT_OBJECTS[py_table_name] = []
            ThisTable.__recid_max__ = 0
            for row in wb_sheet.iter_rows():
                recid = list(row)[0].row
                if ThisTable.__recid_max__ < recid:
                   ThisTable.__recid_max__ = recid
                rec_obj = ThisTable()
                rec_obj.addidx = -1
                if self.has_header:
                    rec_obj.__header_back_map__ = header_back_map
                rec_obj.__recid__ = recid
                rec_obj.__table_name__ += f'[{filename}]{sheet}_{recid}'
                rec_obj.__touched_annotations__ = set()
                self.objects[py_table_name][recid] = rec_obj
                self.mod.HCT_OBJECTS[py_table_name].append(rec_obj)
                sheet_name = hyperc.xtj.str_to_py(f"{sheet}") + f'_{recid}'
                if not hasattr(self.mod.DATA, sheet_name):
                    setattr(self.mod.DATA, sheet_name, self.objects[py_table_name][recid])
                    self.mod.StaticObject.__annotations__[sheet_name] = self.classes[py_table_name]
                
                rec_obj.__py_sheet_name__ = sheet_name

                for _cell in row:
                    xl_orig_calculated_value = getattr(_cell, "value", None)
                    if xl_orig_calculated_value is None:
                        continue
                    letter = _cell.column_letter
                    if is_header:
                        # if xl_orig_calculated_value is None:
                        #     continue
                        header_map[letter] = hyperc.xtj.str_to_py(xl_orig_calculated_value)
                        header_back_map[hyperc.xtj.str_to_py(xl_orig_calculated_value)] = letter
                        continue
                    cell = hyper_etable.cell_resolver.PlainCell(filename=filename, sheet=sheet, letter=letter, number=recid)
                    if self.has_header:
                        column_name = header_map.get(letter, None)
                        #Skip column with empty header bug #176
                        if column_name is None or column_name == "":
                            continue
                    else:
                        column_name = letter
                    if self.has_header:
                        self.objects[py_table_name][recid].__header_back_map__ = header_back_map

                    self.objects[py_table_name][recid].__touched_annotations__.add(column_name)

                    if xl_orig_calculated_value in ['#NAME?', '#VALUE!']:
                        raise Exception(f"We don't support table with error cell {cell}")
                    if (type(xl_orig_calculated_value) == bool or type(xl_orig_calculated_value) == int or type(xl_orig_calculated_value) == str):
                        setattr(self.objects[py_table_name][recid], column_name, xl_orig_calculated_value)
                        setattr(self.objects[py_table_name][recid], column_name, xl_orig_calculated_value)
                        self.objects[py_table_name][recid].__class__.__annotations__[column_name] = str
                        self.objects[py_table_name][recid].__touched_annotations__.add(column_name) 
                    else:
                        setattr(self.objects[py_table_name][recid], column_name, '')
                        self.objects[py_table_name][recid].__class__.__annotations__[column_name] = str
                        self.objects[py_table_name][recid].__touched_annotations__.add(column_name)
                if is_header:
                    is_header = False
                    continue

                for column_name in header_back_map.keys():
                    if not hasattr(rec_obj, column_name):
                        setattr(self.objects[py_table_name][recid], column_name, '')
                        self.objects[py_table_name][recid].__class__.__annotations__[column_name] = str
                        self.objects[py_table_name][recid].__touched_annotations__.add(column_name)

        self.load_external_classes(external_classes_filename)
        for py_table in self.mod.HCT_OBJECTS.values():
            for row in py_table:
                for attr, value in row.__default_init__.items():
                    if value == 'set()':
                       setattr(row, attr, set())
                    elif isinstance(value, str):
                        setattr(row, attr, value.strip('"'))
                    else:
                        setattr(row, attr, value)

        for clsv in self.classes.values():
            var_global_addidx_name = f'DATA.{clsv.__table_name__}_addidx'
            setattr(self.mod.DATA, f'{clsv.__table_name__}_addidx', 0)
            self.mod.StaticObject.__annotations__[f'{clsv.__table_name__}_addidx'] = int
            init_f_code = []
            init_pars = []
            if hyperc.settings.DEBUG:
                print(f" {clsv} -  {clsv.__annotations__}")
            # init_f_code.append(f"global DATA")
            init_f_code.append(f"self.addidx = {var_global_addidx_name}")
            init_f_code.append(f"{var_global_addidx_name} += 1")
            for par_name, par_type in clsv.__annotations__.items():
                if par_name in ['__table_name__', 'addidx']:
                    continue
                # Skip None type cell
                if par_type is None:
                    par_type = str
                    clsv.__annotations__[par_name] = str
                # init_f_code.append(
                    # f'self.{par_name}_not_hasattr = True')  # TODO: statically set to true instead of asking in parameters
                if par_name in clsv.__default_init__:
                    if clsv.__default_init__[par_name] == 'set()':
                        init_f_code.append(f'self.{par_name} = set() # set init "{par_name.upper()}" of table "{clsv.__table_name__}"')
                    else:
                        init_pars.append(f"hct_p_{par_name}:{par_type.__name__}={clsv.__default_init__[par_name]}")
                        init_f_code.append(f'self.{par_name} = hct_p_{par_name} # cell "{par_name.upper()}" of table "{clsv.__table_name__}"')
                else:
                    if not par_type in hyperc.xtj.DEFAULT_VALS:
                        raise TypeError(f"Could not resolve type for {clsv.__name__}.{par_name} (forgot to init cell?)")
                    init_pars.append(f"hct_p_{par_name}:{par_type.__name__}={hyperc.xtj.DEFAULT_VALS[par_type]}")
                    init_f_code.append(f'self.{par_name} = hct_p_{par_name} # cell "{par_name.upper()}" of table "{clsv.__table_name__}"')

            if len(init_f_code) == 0:
                continue
            init_f_code.append(f'side_effect(lambda: HCT_OBJECTS["{clsv.__table_name__}"].append(self))')
            init_f_code.append(f'side_effect(lambda: setattr(self, "__recid__", self.__class__.__recid_max__ + self.addidx))')
            init_f_code.append(f'side_effect(lambda: setattr(self, "__header_back_map__",  self.__class__.__header_back_map__))')
            init_f_code.append(f'side_effect(lambda: setattr(self, "__touched_annotations__",  set()))')
            init_f_code.append(f'side_effect(lambda: [self.__touched_annotations__.add(o) for o in self.__annotations__ if (not (o.startswith("__") and o.endswith("__")) and (o not in getattr(self.__class__,"__user_defined_annotations__", [])) and o != "addidx")])')
            c=f'side_effect(lambda: setattr(DATA, f"{clsv.__table_name__}_'
            init_f_code.append(c+'{self.__recid__}", self))')
            full_f_code = '\n    '.join(init_f_code)
            full_f_pars = ",".join(init_pars)
            full_code = f"def hct_f_init(self, {full_f_pars}):\n    {full_f_code}"
            
            fn = f"{self.tempdir}/hpy_init_{clsv.__name__}.py"
            open(fn, "w+").write(full_code)
            f_code = compile(full_code, fn, 'exec')
            exec(f_code, globals())
            clsv.__init__ = globals()["hct_f_init"]
            clsv.__init__.__name__ = "__init__"

        # Now generate init for static object
        self.mod.DATA.GOAL = False
        self.mod.StaticObject.__annotations__['GOAL'] = bool
        init_f_code = []
        for attr_name, attr_type in self.mod.StaticObject.__annotations__.items():
            init_f_code.append(f"self.{attr_name} = DATA.{attr_name}")  # if it does not ignore, fix it!
        self.mod.StaticObject.__annotations__['GOAL'] = bool
        if init_f_code:

            full_f_code = '\n    '.join(init_f_code)
            full_code = f"def hct_stf_init(self):\n    {full_f_code}"
            fn = f"{self.tempdir}/hpy_stf_init_{self.mod.StaticObject.__name__}.py"
            open(fn, "w+").write(full_code)
            f_code = compile(full_code, fn, 'exec')
            exec(f_code, self.mod.__dict__)
            self.mod.StaticObject.__init__ = self.mod.__dict__["hct_stf_init"]
            self.mod.StaticObject.__init__.__name__ = "__init__"

        # dump goals and actions
        self.dump_functions(goal_code_source, 'hpy_goals.py')

        # addition python code
        for code_file in addition_python_files:
            addition_code = open(code_file, "r").read()
            if addition_code.startswith("from"):  # workaround for module imports
                addition_code = "#"+addition_code
            f_code = compile(addition_code, code_file, 'exec')
            exec(f_code, self.mod.__dict__)
            for f_name in f_code.co_names:
                if "." in f_name: continue  # workaround for module names
                t = self.mod.__dict__.get(f_name, None)
                if isinstance(t, types.FunctionType) or isinstance(t, types.MethodType) or isinstance(t, type):
                    self.methods_classes[f_name] = self.mod.__dict__[f_name]
        # for f in self.mod.__dict__:
        #     if isinstance(self.mod.__dict__[f], types.FunctionType):
        #         self.methods_classes[f] = self.mod.__dict__[f]

        self.methods_classes.update(self.classes)

    def load_external_classes(self, class_py_filename):
        if class_py_filename is None:
            return
        code_str= open(class_py_filename, "r").read()
        code_list = code_str.split("\n")
        code_ast = ast.parse(code_str, filename=class_py_filename, type_comments=True)
        for cl in code_ast.body:
            if not isinstance(cl, ast.ClassDef):
                continue
            if class_in_mod := getattr(self.mod, cl.name, None):
                if hasattr(class_in_mod, '__table_name__'):
                    for ast_obj in cl.body:
                        if f'#{hyper_etable.pysourcebuilder.DEFAULT_COMMENT}' in code_list[ast_obj.lineno-1]:
                            continue
                        if isinstance(ast_obj, ast.AnnAssign):                          
                            class_in_mod.__annotations__[ast_obj.target.id] = pydoc.locate(ast_obj.annotation.id)
                            class_in_mod.__user_defined_annotations__.append(ast_obj.target.id)
                        elif isinstance(ast_obj, ast.FunctionDef) and ast_obj.name =='__init__':
                            for init_line in ast_obj.body:
                                if (f'#{hyper_etable.pysourcebuilder.DEFAULT_COMMENT}' not in code_list[init_line.lineno-1] 
                                    and len(init_line.targets) == 1
                                    and isinstance(init_line.targets[0], ast.Attribute)
                                    and isinstance(init_line.targets[0].value,ast.Name)
                                    and init_line.targets[0].value.id == 'self'):
                                    if isinstance(init_line.value, ast.Constant):
                                        if isinstance(init_line.value.value, str):
                                            value = f'"{init_line.value.value}"'
                                        else:
                                            value = init_line.value.value
                                        class_in_mod.__default_init__[init_line.targets[0].attr] = value
                                    elif isinstance(init_line.value, ast.Call):
                                        if init_line.value.func.id == 'set' and len(init_line.value.args) == 0 :
                                            class_in_mod.__default_init__[init_line.targets[0].attr] = 'set()'  

    def load_rows_in_table(self):
        for obj in self.metadata['new_instances']:
            if hasattr(obj, '__table_name__') and hasattr(obj, 'addidx') :
                obj.__recid__ = obj.addidx + obj.__recid_max__ + 1
                if hasattr(obj.__class__, '__header_back_map__'):
                    obj.__header_back_map__ = obj.__class__.__header_back_map__
                self.mod.HCT_OBJECTS[obj.__table_name__].append(obj)
                setattr(self.mod.DATA,f'{obj.__table_name__}_{obj.__recid__}', obj)
                for ann in [o for o in obj.__annotations__ if not (o.startswith('__') and o.endswith('__'))]:
                    if ann in getattr(obj.__class__,'__user_defined_annotations__', []):
                        continue
                    if ann == 'addidx':
                        continue
                    obj.__touched_annotations__.add(ann)

    def reset_data(self):
        for table in self.mod.HCT_OBJECTS.values():
            for row in table:
                sheet_name = row.__xl_sheet_name__
                recid = row.__recid__
                for attr_name in row.__touched_annotations__:
                    if self.has_header:
                        letter = row.__header_back_map__[attr_name]
                    else:
                        letter = attr_name
                    if getattr(self.wb_values_only[sheet_name][f'{letter}{recid}'], "value", None) is None:
                        continue
                    old_value = self.wb_values_only[sheet_name][f'{letter}{recid}'].value
                    setattr(row, attr_name, old_value)
        for obj in self.metadata['new_instances']:
            if hasattr(obj, '__table_name__') and hasattr(obj, 'addidx') :
                deleted = True
                for table in self.mod.HCT_OBJECTS.values():
                    if obj in table:
                        table.remove(obj)
                        break
            
                if deleted:
                    delattr(self.mod.DATA,f'{obj.__table_name__}_{obj.__recid__}')    
        self.metadata['new_instances'] = []
        self.mod.DATA.GOAL = False

    def generate_invariants(self):
        def gg(s, g, e):
            return hyperc.solve(g, globals_=s.methods_classes, extra_instantiations=e, work_dir=s.tempdir, 
                            addition_modules=[s.mod], metadata=s.metadata)
        self.metadata = {"GENERATE_INVARIANTS": []}
        invariants = gg(self,self.methods_classes[self.main_goal.name],
                                              list(filter(lambda x: isinstance(x, type), self.methods_classes.values())))
        return invariants

    def dump_py(self, dir=None, out_filename=None):
        """"Dump classes as python code"""
        if dir is None:
            dir =  self.filename.parent
        try:
            os.mkdir(dir)
        except FileExistsError:
            pass 
        
        # dump classes as python code
        for c in itertools.chain([hyper_etable.meta_table.TableElementMeta], self.classes.values(), [self.mod.StaticObject, self.mod.DefinedTables]):
            self.source_code['classes'].append(hyper_etable.pysourcebuilder.build_source_from_class(c, ['__table_name__','__xl_sheet_name__'], default_comment=hyper_etable.pysourcebuilder.DEFAULT_COMMENT).end())

        # dump object as python code
        self.source_code['classes'].append('DATA = StaticObject()')
        if out_filename is None:
            for f_name, code  in self.source_code.items():
                code_file = os.path.join(dir, f'{f_name}.py')
                s_code =""
                for func in code:
                    s_code += str(func)
                    s_code += '\n'
                open(code_file, "w+").write(s_code)
        else:
            s_code =""
            for f_name, code  in self.source_code.items():
                for func in code:
                    s_code += str(func)
                    s_code += '\n'
            open(out_filename, "w+").write(s_code)

    def run_plan(self, py_plan_filename):
        """Run python plan"""
        plan_code_str = open(py_plan_filename, "r").read()
        f_code = compile(plan_code_str, py_plan_filename, 'exec')
        exec(f_code, self.mod.__dict__)

    def save_plan(self, prefix="DATA.", exec_plan=False, out_dir=None, out_filename=None):
        """Dump plan as python code"""
        self.plan_data_prefix=prefix
        if out_dir is None:
            out_dir =  os.path.join(self.filename.parent, 'out')
        if out_filename is None:
            self.plan_file = pathlib.Path(os.path.join(out_dir,f'{os.path.splitext(self.filename.name)[0]}.py'))
        else:
            self.plan_file = out_filename
        try:
            os.mkdir(out_dir)
        except FileExistsError:
            pass 
        code = []
        for step in self.metadata["plan_exec"]:
            args = []
            i=0
            for k, a in step[1].items():
                if hasattr(a,"__py_sheet_name__"):
                    args.append(f'{k}={prefix}{a.__py_sheet_name__}')
                else:
                    args.append(f'{k}=unresolved_variable_{i}')
            args = ", ".join(args)
            code.append(f'{step[0].__name__}({args})')
        code_str = "\n".join(code)
        with open(self.plan_file, "w+") as f:
            f.write(code_str)
        if exec_plan:
            self.run_plan(py_plan_filename=self.plan_file)

    def save_dump(self, has_header=False, out_dir=None, out_filename=None):
        """Save objects into XLSX file"""
        if out_dir is None:
            out_dir =  os.path.join(self.filename.parent, 'out')
        try:
            os.mkdir(out_dir)
        except FileExistsError:
            pass
        for table in self.mod.HCT_OBJECTS.values():
            for row in table:
                sheet_name = row.__xl_sheet_name__
                recid = row.__recid__
                for attr_name in row.__touched_annotations__:
                    if self.has_header:
                        letter = row.__header_back_map__[attr_name]
                    else:
                        letter = attr_name
                    new_value = getattr(row, attr_name)
                    if getattr(self.wb_values_only[sheet_name][f'{letter}{recid}'], "value", None) == new_value:
                        continue
                    self.wb_values_only[sheet_name][f'{letter}{recid}'].value = new_value
                    self.wb_with_formulas[sheet_name][f'{letter}{recid}'].value = new_value
        if out_filename is None:
            out_filename = os.path.join(out_dir, f'{self.filename.name}')
        self.wb_with_formulas.save(out_filename)
        return out_filename

    def calculate(self):

        # g=self.get_range_name_by_cell("'[fff]ggg'!B1")
        # gg = hyper_etable.etable_transpiler.split_cell(
        #     "ONESTABLE[PLUS ONES]")
        self.range_resolver = hyper_etable.cell_resolver.RangeResolver(os.path.basename(self.filename), self.wb_with_formulas)

        xl_mdl = formulas.excel.ExcelModel()
        xl_mdl.loads(str(self.filename))
        # 
        self.stl = hyper_etable.spiletrancer.SpileTrancer(self.filename, xl_mdl, self.mod.DATA, plan_log=self.plan_log)
        var_mapper = {}
        global_table_type_mapper = {}
        code = {}
        filename_case_remap_workaround = {}

        used_cell_set = set()

        for ws in self.wb_with_formulas.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue # skip empty cell
                    text_formula = str(cell.value)
                    current_cell = hyper_etable.cell_resolver.PlainCell(
                        filename=self.filename, sheet=ws.title, letter=cell.column_letter, number=cell.row)
                    if not text_formula.startswith("="):
                        self.cells_value[current_cell] = cell.value
                        continue # pass only formulas
                    output = hyper_etable.etable_transpiler.StringLikeVariable(
                        var_map=var_mapper, filename=self.filename, sheet=ws.title, letter=cell.column_letter, number=cell.row)
                    out_py = hyperc.xtj.str_to_py(f'[{output.filename}]{output.sheet}!{output.letter}{output.number}')
                    code_init = hyper_etable.etable_transpiler.FunctionCode(name=f'hct_{out_py}')
                    code_init.init.append(f'#{text_formula}')

                    used_cell_set.add(current_cell)
                    formula = hyper_etable.etable_transpiler.EtableTranspiler(
                        formula=text_formula, range_resolver= self.range_resolver,
                        output=output, init_code=code_init, table_type_mapper=global_table_type_mapper, var_mapper=var_mapper)
                    formula.transpile_start()
                    # set default value for takeif
                    var = formula.default
                    if isinstance(formula.default, hyper_etable.etable_transpiler.StringLikeConstant):
                        var = formula.default.var
                    self.cells_value[current_cell] = var
                    code.update(formula.code)

        # Find what is being watched, and inject watchtakeif sync cells
        watched_takeifs = defaultdict(list)
        for func in code.values():
            if func.watchtakeif:
                watched_takeifs[func.watchtakeif].append(func)
        min_recid = {}
        max_recid = {}
        for watched_takeif_cell, watched_actions in  watched_takeifs.items():
            for w_a in watched_actions:
                cell = list(w_a.effect_vars)[0]
                watch_for = f'{watched_takeif_cell}_{cell.letter}'
                if watch_for not in min_recid:
                    min_recid[watch_for] = cell
                    max_recid[watch_for] = cell
                if cell.number > max_recid[watch_for].number:
                    max_recid[watch_for] = cell
                if cell.number < max_recid[watch_for].number:
                    min_recid[watch_for] = cell

        watch_code = []
        for watch_for, recid in min_recid.items():
            watch_code.append(f"WATCHTAKEIF_{watch_for} = {recid.number}")
            watch_code.append(f"WATCHTAKEIF_MAX_{watch_for} = {max_recid[watch_for].number+1}")
        if len(watch_code) > 0:
            watch_code = "\n".join(watch_code)
            fn=f"{self.tempdir}/hpy_watch_code.py"
            with open(fn, "w+") as f:
                f.write(watch_code)
            f_code=compile(watch_code, fn, 'exec')
            exec(f_code, self.mod.__dict__)

        for func in code.values():
            used_cell_set.update(set([i.cell for i in func.input_variables]))
            for var in func.sync_cell:
                if not isinstance(var, hyper_etable.etable_transpiler.StringLikeVariable):
                    continue
                cell_name = var.get_excel_format()
                if (cell_name in used_cell_set) and (cell_name not in xl_mdl.cells):
                    used_cell_set.remove(cell_name)

        # # look for mergable actions by sync
        deleted_keys = set()
        for func_name_other in list(code.keys()):
            if func_name_other in deleted_keys:
                continue
            for func_name in list(code.keys()):
                #check that funtions is not parent and child
                if func_name_other in deleted_keys:
                    continue
                if code[func_name] is code[func_name_other]:
                    continue
                if (not code[func_name].sync_cell.isdisjoint(code[func_name_other].sync_cell)):
                    code[func_name_other].merge(code[func_name])
                    del code[func_name]
                    deleted_keys.add(func_name)

        # merge watchtakeif's
        deleted_keys = set()
        for watchif_func_name in list(code.keys()):
            if watchif_func_name in deleted_keys:
                continue
            if code[watchif_func_name].watchtakeif is None:
                continue
            for watchif_func_name_other in list(code.keys()):
                if code[watchif_func_name_other].watchtakeif is None:
                    continue
                if watchif_func_name_other == watchif_func_name:
                    continue
                if code[watchif_func_name_other].watchtakeif == code[watchif_func_name].watchtakeif:
                    code[watchif_func_name].merge(code[watchif_func_name_other])
                    del code[watchif_func_name_other]
                    deleted_keys.add(watchif_func_name_other)

        deleted_keys = set()
        for watchif_func_name in list(code.keys()):
            deleted = False
            for func_name in list(code.keys()):
                if code[func_name].watchtakeif is not None:
                    continue
                if code[watchif_func_name].watchtakeif is None:
                    continue
                if code[watchif_func_name].watchtakeif in code[func_name].effect_vars:
                    code[func_name].merge(code[watchif_func_name])
                    deleted = True
            if deleted:
                del code[watchif_func_name]

        # update keys
        code = {v.name: v for k, v in code.items()}


        # Collect conditional formatting
        # TODO set goal here
        goal_code = defaultdict(list)
        goal_code_used_vars = set()
        for filename, book in xl_mdl.books.items():
            filename = filename_case_remap_workaround.get(filename, filename)
            for worksheet in book[formulas.excel.BOOK].worksheets:
                for rule_cell in worksheet.conditional_formatting._cf_rules:
                    assert len(rule_cell.sqref.ranges) == 1, "Only one cell conditional rule is supported"
                    sheet = worksheet.title
                    assert rule_cell.sqref.ranges[0].size == {'columns': 1, 'rows': 1}, "Ranged rules are not supported"
                    cell = f"'[{filename}]{sheet}'!{rule_cell.sqref.ranges[0].coord}"
                    col_letter = hyper_etable.util.get_char_from_index(rule_cell.sqref.ranges[0].bounds[0]-1)
                    row = rule_cell.sqref.ranges[0].bounds[1]
                    current_cell = hyper_etable.cell_resolver.PlainCell(
                        filename=filename, sheet=sheet, letter=col_letter, number=row)
                    used_cell_set.add(current_cell)
                    goal_code_used_vars.add(current_cell)
                    filename_, sheet, recid, letter = hyper_etable.etable_transpiler.split_cell(cell)
                    sheet_name = hyperc.xtj.str_to_py(f"[{filename}]{sheet}") + f'_{recid}'
                    for rule in rule_cell.rules:
                        value = hyper_etable.etable_transpiler.formulas_parser(rule.formula[0])[0]
                        if isinstance(value, formulas.tokens.operand.Range):
                            filename_value, sheet_value, recid_value, letter_value = hyper_etable.etable_transpiler.split_cell(
                                rule.formula[0])
                            # if filename_value == '':
                            filename_value = filename
                            if sheet_value == '':
                                sheet_value = sheet
                            # used_cell_set.add(f"'[{filename_value}]{sheet_value}'!{letter_value.upper()}{recid_value}")
                            current_cell = hyper_etable.cell_resolver.PlainCell(
                                filename=filename_value, sheet=sheet_value, letter=letter_value.upper(), number=recid_value)
                            goal_code_used_vars.add(current_cell)
                            used_cell_set.add(current_cell)
                            sheet_name_value = hyperc.xtj.str_to_py(
                                f"[{filename_value}]{sheet_value}") + f'_{recid_value}'
                            goal_code[cell].append([
                                f'assert DATA.{sheet_name}.{letter} {operator_name_to_operator(rule.operator)} DATA.{sheet_name_value}.{letter_value}',
                                f'assert DATA.{sheet_name}.{letter}_not_hasattr == False', 
                                f'assert DATA.{sheet_name_value}.{letter_value}_not_hasattr == False'])
                        elif isinstance(value, formulas.tokens.operand.Number):
                            if str(value.attr["name"]) == "TRUE":
                                goal_code[cell].append([
                                    f'assert DATA.{sheet_name}.{letter} {operator_name_to_operator(rule.operator)} True',
                                    f'assert DATA.{sheet_name}.{letter}_not_hasattr == False'])
                            elif str(value.attr["name"]) == "FALSE":
                                goal_code[cell].append([
                                    f'assert DATA.{sheet_name}.{letter} {operator_name_to_operator(rule.operator)} False',
                                    f'assert DATA.{sheet_name}.{letter}_not_hasattr == False'])
                            else:
                                goal_code[cell].append([
                                    f'assert DATA.{sheet_name}.{letter} {operator_name_to_operator(rule.operator)} {int(value.attr["name"])}',
                                    f'assert DATA.{sheet_name}.{letter}_not_hasattr == False'])
                        elif isinstance(value, formulas.tokens.operand.String):
                            goal_code[cell].append([
                                f'assert DATA.{sheet_name}.{letter} {operator_name_to_operator(rule.operator)} "{value.attr["name"]}"',
                                f'assert DATA.{sheet_name}.{letter}_not_hasattr == False'])

        g_c = hyper_etable.etable_transpiler.FunctionCode(name='condition_goal', is_goal=True)
        goal_code_source = {}
        goal_code_source[0] = hyper_etable.etable_transpiler.FunctionCode(name=f'hct_goal_0', is_goal=True)
        goal_code_source[0].output[goal_code_source[0].name].append('DATA.GOAL = True')
        goal_code_source[0].output[goal_code_source[0].name].append('pass')

        goal_counter = 0
        for goal_name, g_c in goal_code.items():
            goal_counter_was = goal_counter
            goal_counter = (goal_counter + 1) * len(g_c) - 1
            if goal_counter > goal_counter_was:
                counter_was = 0
                for counter_new in range(goal_counter_was + 1, goal_counter+1):
                    if counter_was == goal_counter_was + 1:
                        counter_was = 0
                    goal_code_source[counter_new] = hyper_etable.etable_transpiler.FunctionCode(
                        name=f'hct_goal_{counter_new}', is_goal=True)
                    goal_code_source[counter_new].operators[goal_code_source[counter_new].name] = copy.copy(goal_code_source[counter_was].operators)
                    goal_code_source[counter_new].output[goal_code_source[counter_new].name].append(
                        'DATA.GOAL = True')
                    goal_code_source[counter_new].input_variables.update(goal_code_used_vars)
                    goal_code_source[counter_new].all_variables.update(goal_code_used_vars)
                    counter_was += 1

            for idx in goal_code_source:
                goal_code_source[idx].operators[goal_code_source[idx].name].append(f'#{goal_name}')
                goal_code_source[idx].operators[goal_code_source[idx].name].extend(g_c[idx % len(g_c)])


        main_goal = hyper_etable.etable_transpiler.FunctionCode(name=f'hct_main_goal', is_goal=True)
        main_goal.operators[main_goal.name].append('assert DATA.GOAL == True')
        main_goal.operators[main_goal.name].append('pass')
        goal_code_source['main_goal'] = main_goal

        # delete tailing actions
        unused_cell_set = set()
        some_found = True
        while some_found: #double pass search
            some_found = False
            for function_key_deletable in list(code.keys()):
                func_deletable = code.get(function_key_deletable, None)
                if func_deletable is None:
                    continue
                found = False
                effect_vars = hyper_etable.etable_transpiler.unpack_cell(func_deletable.effect_vars)
                # look in actions
                for function_key in list(code.keys()):
                    func = code.get(function_key, None)
                    if func is None:
                        continue
                    input_variables = hyper_etable.etable_transpiler.unpack_cell(func.input_variables)
                    if effect_vars & input_variables:
                        found = True
                        break
                if found:
                    continue
                # look in goals
                if effect_vars & goal_code_used_vars:
                    found = True
                    continue
                if not found:
                    unused_cell_set.update(effect_vars)
                    del code[function_key_deletable]
                    some_found = True
        # Generate init code
        [c.gen_init() for c in code.values()]

        # Load used cell
        for cell in used_cell_set:
            filename = cell.filename
            sheet = cell.sheet
            recid_ret = cell.number

            letter_ret = cell.letter

            filename = filename_case_remap_workaround.get(filename, filename)
            py_table_name = hyperc.xtj.str_to_py(f'[{filename}]{sheet}')
            if isinstance(recid_ret, list):
                recid_ret = range(recid_ret[0], recid_ret[1] + 1)
            else:
                recid_ret = [recid_ret]

            if isinstance(letter_ret, list):
                letter_stop = letter_ret[1]
                letter_next = letter_ret[0]
                letter_ret = [letter_next]
                while letter_next != letter_stop:
                    letter_next = hyperc.util.letter_index_next(letter = letter_next).upper()
                    letter_ret.append(letter_next)
            else:
                letter_ret = [letter_ret]
            for letter in letter_ret:
                for recid in recid_ret:
                    cell = hyper_etable.cell_resolver.PlainCell(filename=filename, sheet=sheet, letter=letter, number=recid)
                    if cell in unused_cell_set:
                        continue
                    if recid not in self.objects[py_table_name]:
                        if py_table_name not in self.classes:
                            ThisTable = self.get_new_table(py_table_name, sheet)
                        else:
                            ThisTable = self.classes[py_table_name]
                        # if 
                        rec_obj = ThisTable()
                        # rec_obj.__row_record__ = copy.copy(cell)
                        rec_obj.__table_name__ += f'[{filename}]{sheet}_{recid}'
                        rec_obj.__touched_annotations__ = set()
                        # ThisTable.__annotations_type_set__ = defaultdict(set)
                        self.objects[py_table_name][recid] = rec_obj
                        self.mod.HCT_OBJECTS[py_table_name].append(rec_obj)
                    
                    # Declare and load defined table names
                    defined_table_name = self.range_resolver.get_table_by_cell(cell)
                    if defined_table_name is not None:
                        for dtn_raw in defined_table_name:
                            dtn = hyperc.xtj.str_to_py(dtn_raw)
                            if dtn not in self.mod.DefinedTables.__annotations__:
                                self.mod.DefinedTables.__annotations__[dtn] = set
                                setattr(self.mod.DEFINED_TABLES, dtn, set())
                            getattr(self.mod.DEFINED_TABLES, dtn).add(self.objects[py_table_name][recid])

                    self.objects[py_table_name][recid].__touched_annotations__.add(letter)
                    self.objects[py_table_name][recid].__annotations__[(f'{letter}_not_hasattr')] = bool
                    #TODO add type detector
                    # self.classes[py_table_name].__annotations__[letter] = int
                    # rec_obj.__annotations__.add(letter)
                    sheet_name = hyperc.xtj.str_to_py(f"[{filename}]{sheet}") + f'_{recid}'
                    if not hasattr(self.mod.DATA, sheet_name):
                        setattr(self.mod.DATA, sheet_name, self.objects[py_table_name][recid])
                        self.mod.StaticObject.__annotations__[sheet_name] = self.classes[py_table_name]
                    # assert cell in self.cells_value, f"Lost value for cell {cell}"
                    if cell not in self.cells_value or self.cells_value[cell] is None:
                        # TODO this is stumb for novalue cell. We should use Novalue ????
                        ox_sht, ox_cell_ref = self.stl.gen_opxl_addr(self.filename, 
                                                        self.objects[py_table_name][recid].__class__.__xl_sheet_name__, 
                                                        letter, recid)
                        xl_orig_calculated_value = self.wb_values_only[ox_sht][ox_cell_ref].value
                        if xl_orig_calculated_value in ['#NAME?', '#VALUE!']  and self.enable_precalculation:
                            raise Exception("We don't support table with error cell")
                        elif not self.enable_precalculation:
                            xl_orig_calculated_value = ''
                        if (type(xl_orig_calculated_value) == bool or type(xl_orig_calculated_value) == int or type(xl_orig_calculated_value) == str) and self.enable_precalculation:
                            setattr(self.objects[py_table_name][recid], letter, xl_orig_calculated_value)
                            self.objects[py_table_name][recid].__class__.__annotations__[letter] = str
                        else:
                            setattr(self.objects[py_table_name][recid], letter, '')
                            self.objects[py_table_name][recid].__class__.__annotations__[letter] = str
                        setattr(self.objects[py_table_name][recid], f'{letter}_not_hasattr', True)

                    else:
                        setattr(self.objects[py_table_name][recid], letter, self.cells_value[cell])
                        setattr(self.objects[py_table_name][recid], f'{letter}_not_hasattr', False)
                        # FIXME: needs type detector, then these lines can be removed -->
                        # self.objects[py_table_name][recid].__class__.__annotations__[letter] = type(cell_value)
                        # self.objects[py_table_name][recid].__annotations__[letter] = type(cell_value)
                        self.objects[py_table_name][recid].__class__.__annotations__[letter] = str  # bug hyperc#453
                        self.objects[py_table_name][recid].__annotations__[letter] = str  # bug hyperc#453 
                        # <-- end FIXME
        # Type detector
        # Match all group neighbor each other
        # by Breadth-first search now
        # not_double_pass = True
        # while not_double_pass:
        #     not_double_pass = False
        #     for tm in global_table_type_mapper.values():
        #         while tm.visited_group != tm.group:
        #             tm.visited_group = tm.group
        #             tmp_tm_group = copy.copy(tm.group)
        #             for tm_name in tmp_tm_group:
        #                 tm.merge_group(global_table_type_mapper[tm_name])
        #     for tm in global_table_type_mapper.values():
        #         while tm.forward_visited_group != tm.group:
        #             tm.forward_visited_group = tm.group
        #             tmp_tm_group = copy.copy(tm.group)
        #             for tm_name in tmp_tm_group:
        #                 tm.merge_group(global_table_type_mapper[tm_name])
        #         if tm.forward_visited_group != tm.visited_group:
        #             not_double_pass = True

        # for var in var_mapper.values():
        #     if isinstance(var, hyper_etable.etable_transpiler.StringLikeVariable):
        #         if var.is_range:
        #             recid_ret = range(var.number[0], var.number[1]+1)
        #             letter = var.letter[0]
        #         else:
        #             recid_ret = [var.number]
        #             letter = var.letter
        #         py_table_name = hyperc.xtj.str_to_py(f'[{var.filename}]{var.sheet}')
        #         for recid in recid_ret:
        #             line_object = self.objects[py_table_name][recid]
        #             #TODO fix type detector for ranges
        #             # if int in var.types:
        #                 # line_object.__annotations__[letter] = int
        #                 # line_object.__class__.__annotations__[letter] = int
        #             line_object.__annotations__[letter] = int
        #             line_object.__class__.__annotations__[letter] = int

        # stack_code = ''
        # for cell in used_cell_set:
        #     filename, sheet, recid_ret, letter = hyper_etable.etable_transpiler.split_cell(cell)
        #     stack_code = stack_code_gen(hyperc.xtj.str_to_py(f'[{filename}]{sheet}'))
        #     break

        # Dump defined table names
        init_f_code = []
        for attr_name, attr_type in self.mod.DefinedTables.__annotations__.items():
            init_f_code.append(f"self.{attr_name} = DEFINED_TABLES.{attr_name}")  # if it does not ignore, fix it!
        self.mod.DefinedTables.__annotations__['GOAL'] = bool
        if init_f_code:
            full_f_code = '\n    '.join(init_f_code)
            full_code = f"def hct_dt_init(self):\n    {full_f_code}"
            fn = f"{self.tempdir}/hpy_dt_init.py"
            open(fn, "w+").write(full_code)
            f_code = compile(full_code, fn, 'exec')
            exec(f_code, self.mod.__dict__)
            self.mod.DefinedTables.__init__ = self.mod.__dict__["hct_dt_init"]
            self.mod.DefinedTables.__init__.__name__ = "__init__"


        stack_code = stack_code_gen_all(self.objects)

        fn = f"{self.tempdir}/hpy_stack_code.py"
        with open(fn, "w+") as f:
            f.write(str(stack_code))

        f_code = compile(stack_code, fn, 'exec')
        exec(f_code, self.mod.__dict__)

        for clsv in self.classes.values():
            init_f_code = []
            init_pars = []
            if hyperc.settings.DEBUG:
                print(f" {clsv} -  {clsv.__annotations__}")
            for par_name, par_type in clsv.__annotations__.items():
                if par_name == '__table_name__':
                    continue
                # Skip None type cell
                if par_type is None:
                    par_type = str
                    clsv.__annotations__[par_name] = str
                init_f_code.append(
                    f'self.{par_name} = hct_p_{par_name} # cell "{par_name.upper()}" of table "{clsv.__table_name__}"')
                # init_f_code.append(
                    # f'self.{par_name}_not_hasattr = True')  # TODO: statically set to true instead of asking in parameters
                if not par_type in hyperc.xtj.DEFAULT_VALS:
                    raise TypeError(f"Could not resolve type for {clsv.__name__}.{par_name} (forgot to init cell?)")
                init_pars.append(f"hct_p_{par_name}:{par_type.__name__}={hyperc.xtj.DEFAULT_VALS[par_type]}")
            if len(init_f_code) == 0:
                continue
            full_f_code = '\n    '.join(init_f_code)
            full_f_pars = ",".join(init_pars)
            full_code = f"def hct_f_init(self, {full_f_pars}):\n    {full_f_code}"
            fn = f"{self.tempdir}/hpy_init_{clsv.__name__}.py"
            open(fn, "w+").write(full_code)
            f_code = compile(full_code, fn, 'exec')
            exec(f_code, globals())
            clsv.__init__ = globals()["hct_f_init"]
            clsv.__init__.__name__ = "__init__"


        # Now generate init for static object
        self.mod.DATA.GOAL = False
        self.mod.StaticObject.__annotations__['GOAL'] = bool
        init_f_code = []
        for attr_name, attr_type in self.mod.StaticObject.__annotations__.items():
            init_f_code.append(f"self.{attr_name} = DATA.{attr_name}")  # if it does not ignore, fix it!
        self.mod.StaticObject.__annotations__['GOAL'] = bool
        if init_f_code:

            full_f_code = '\n    '.join(init_f_code)
            full_code = f"def hct_stf_init(self):\n    {full_f_code}"
            fn = f"{self.tempdir}/hpy_stf_init_{self.mod.StaticObject.__name__}.py"
            open(fn, "w+").write(full_code)
            f_code = compile(full_code, fn, 'exec')
            exec(f_code, self.mod.__dict__)
            self.mod.StaticObject.__init__ = self.mod.__dict__["hct_stf_init"]
            self.mod.StaticObject.__init__.__name__ = "__init__"



        #dump goals and actions
        self.dump_functions(code, 'hpy_etable.py')
        self.dump_functions(goal_code_source, 'hpy_goals.py')

        self.methods_classes.update(self.classes)
        just_classes = list(filter(lambda x: isinstance(x, type), self.methods_classes.values()))

        # plan_or_invariants = hyperc.solve(self.methods_classes[main_goal.name], self.methods_classes, just_classes, DATA)

        plan_or_invariants = self.solver_call(goal=self.methods_classes[main_goal.name],
                                              extra_instantiations=just_classes)
        print("finish")
        
    def finish(self):
        self.stl.calculate_excel()
        dirn = os.path.dirname(self.filename)
        new_dirname = os.path.join(dirn, f"{self.filename.name}_out")
        try:
            mkdir(new_dirname)
        except FileExistsError:
            pass
        new_dirname_forfile = os.path.join(dirn, f"{self.filename.name}_out", str(int(time.time())))
        mkdir(new_dirname_forfile)
        self.stl.write(new_dirname_forfile)

        self.out_filename = os.path.join(new_dirname_forfile, self.filename.name)
        return self.out_filename

    # def call_sequential(self, sequency)
