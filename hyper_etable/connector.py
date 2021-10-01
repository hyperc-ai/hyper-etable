import hyperc.xtj
import openpyxl
import hyper_etable.meta_table
import hyper_etable.ms_api
import hyperc.util
import collections
import io
import googleapiclient.discovery
import googleapiclient.http
import os.path
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import google.auth.exceptions
from google.oauth2.credentials import Credentials
import openpyxl
import pathlib
import requests
from hyper_etable.util import OrderedSet

def new_connector(path, proto, mod, has_header=True):
    if proto.lower() == 'msapi':
        conn = hyper_etable.connector.MSAPIConnector(path, mod, has_header)
    elif proto.lower() == 'gsheet':
        conn = hyper_etable.connector.GSheetConnector(path, mod, has_header)
    elif proto.lower() == 'xlsx':
        conn = hyper_etable.connector.XLSXConnector(path, mod, has_header)
    elif proto.lower() == 'airtable':
        conn = hyper_etable.connector.AirtableConnector(path, mod, has_header)
    elif proto.lower() == 'mysql':
        conn = hyper_etable.connector.MySQLConnector(path, mod, has_header)
    elif proto.lower() == 'sqlalchemy':
        conn = hyper_etable.connector.SQLAlchemyConnector(path, mod, has_header)
    if conn is None:
        raise ValueError(f'{proto} is not support')
    return conn

class NameMap:
    name: str
    letter: str
    def __init__(self, name, letter):
        self.column_name = name
        self.letter = letter

class Connector:
    def __init__(self, path, mod, has_header=True):
        self.path = path
        self.has_header = has_header
        self.tables = {}
        self.raw_tables = {}
        self.objects = {}
        self.classes = {}
        self.mod = mod

    def get_raw_table(self):
        tables = collections.defaultdict(dict)
        for table_name, table in self.mod.HCT_OBJECTS.items():
            if self.classes[table_name].__connector__ is self:
                for row in table:
                    tables[row.__xl_sheet_name__][row.__recid__] = {}
                    for column_name in row.__touched_annotations__:
                        if hasattr(row.__class__, '__column_to_py_map__'):
                            tables[row.__xl_sheet_name__][row.__recid__][row.__class__.__column_to_py_map__[column_name]] = getattr(row, column_name)
                        elif hasattr(row.__class__, '__header_back_map__') and self.has_header:
                            tables[row.__xl_sheet_name__][row.__recid__][row.__header_back_map__[column_name]] = getattr(row, column_name)
                        else:
                            tables[row.__xl_sheet_name__][row.__recid__][column_name] = getattr(row, column_name)
        return dict(tables)

    def get_all_raw_table(self):
        tables = collections.defaultdict(dict)
        for table_name, table in self.mod.HCT_OBJECTS.items():
            for row in table:
                tables[row.__xl_sheet_name__][row.__recid__] = {}
                for column_name in row.__touched_annotations__:
                    if hasattr(row.__class__, '__column_to_py_map__'):
                        column_name_wrap = NameMap(row.__class__.__column_to_py_map__[column_name], )
                        tables[row.__xl_sheet_name__][row.__recid__][column_name_wrap] = getattr(row, column_name)

                    elif hasattr(row.__class__, '__header_back_map__') and self.has_header:
                        tables[row.__xl_sheet_name__][row.__recid__][row.__header_back_map__[column_name]] = getattr(row, column_name)
                    else:
                        tables[row.__xl_sheet_name__][row.__recid__][column_name] = getattr(row, column_name)
        return tables
    
    def calculate_columns(self):
        tables = {}
        for table_name in self.HCT_OBJECTS.keys():
            if self.classes[table_name].__connector__ is not self:
                continue
            tables[table_name] = self.classes[table_name].__header_back_map__

        return tables
        
    def reload(self):
        self.raw_tables = {}
        self.load()

    def load(self):
        tables_was = set(list(self.raw_tables.keys()))
        raw_tables = self.load_raw()
        self.raw_tables.update(raw_tables)
        table_name_set = set(list(self.raw_tables.keys())) - tables_was
        for table_name in table_name_set:
            self.load_table(table_name)

    def load_table(self, table_name):

        py_table_name = hyperc.xtj.str_to_py(f'{table_name}') # warning only sheet in 
        if py_table_name in self.mod.HCT_OBJECTS:
            raise ValueError(f'Error sheet {table_name} already exist')
        ThisTable = hyper_etable.meta_table.TableElementMeta(f'{py_table_name}_Class', (object,), {'__table_name__': py_table_name, '__xl_sheet_name__': table_name})
        ThisTable.__annotations__ = {'__table_name__': str, 'addidx': int}
        ThisTable.__user_defined_annotations__ = []
        ThisTable.__default_init__ = {}
        ThisTable.__touched_annotations__ = OrderedSet()
        ThisTable.__annotations_type_set__ = collections.defaultdict(set)
        ThisTable.__connector__ = self
        ThisTable.__column_to_py_map__ = {}
        self.mod.__dict__[f'{py_table_name}_Class'] = ThisTable
        self.classes[py_table_name] = ThisTable
        self.classes[py_table_name].__qualname__ = f"{self.mod.__name__}.{py_table_name}_Class"
        self.mod.HCT_OBJECTS[py_table_name] = []
        self.objects[py_table_name]={}
        ThisTable.__recid_max__ = 0
        recid = 0
        for recid, row in self.raw_tables[table_name].items():
            if ThisTable.__recid_max__ < recid:
                ThisTable.__recid_max__ = recid
            rec_obj = ThisTable()
            rec_obj.addidx = -1
            rec_obj.__recid__ = recid
            rec_obj.__table_name__ += f'[mysql]{table_name}_{recid}'
            rec_obj.__touched_annotations__ = OrderedSet()
            rec_obj.__xl_sheet_name__ = table_name
            self.objects[py_table_name][recid] = rec_obj
            self.mod.HCT_OBJECTS[py_table_name].append(rec_obj)
            sheet_name = hyperc.xtj.str_to_py(f"{table_name}") + f'_{recid}'
            if not hasattr(self.mod.DATA, sheet_name):
                setattr(self.mod.DATA, sheet_name, self.objects[py_table_name][recid])
                self.mod.StaticObject.__annotations__[sheet_name] = self.classes[py_table_name]
            
            rec_obj.__py_sheet_name__ = sheet_name

            for column_name, value in self.raw_tables[table_name][recid].items():
                py_column_name = hyperc.xtj.str_to_py(f'{column_name}')
                ThisTable.__column_to_py_map__[py_column_name] = column_name
                self.objects[py_table_name][recid].__touched_annotations__.add(py_column_name)
                if (type(value) == bool or type(value) == int or type(value) == str):
                    setattr(self.objects[py_table_name][recid], py_column_name, value)
                    self.objects[py_table_name][recid].__class__.__annotations__[py_column_name] = str
                    self.objects[py_table_name][recid].__touched_annotations__.add(py_column_name) 
                else:
                    setattr(self.objects[py_table_name][recid], py_column_name, '')
                    self.objects[py_table_name][recid].__class__.__annotations__[py_column_name] = str
                    self.objects[py_table_name][recid].__touched_annotations__.add(py_column_name)
    def save(self):
        raw_tables_to_save = self.get_raw_table()
        raw_tables_in_base = self.load_raw()
        raw_tables_to_update = collections.defaultdict(dict)
        raw_tables_to_append = collections.defaultdict(dict)
        for table_name, table in raw_tables_to_save.items():
            if table_name in raw_tables_in_base: # save tables available in table
                for recid in sorted(table.keys()):
                    if recid in raw_tables_in_base[table_name]:
                        if raw_tables_in_base[table_name][recid] !=  raw_tables_to_save[table_name][recid]:
                            raw_tables_to_update[table_name][recid] = raw_tables_to_save[table_name][recid]
                    else:
                        raw_tables_to_append[table_name][recid] = raw_tables_to_save[table_name][recid]
        self.raw_update(raw_tables_to_update)
        self.raw_append(raw_tables_to_append)

    def save_all(self, raw_tables_in_base = None):
        raw_tables_to_save = self.get_all_raw_table()
        if raw_tables_in_base is not None:
            raw_tables_in_base = self.load()
        raw_tables_to_update = collections.defaultdict(dict)
        raw_tables_to_append = collections.defaultdict(dict)
        for table_name, table in raw_tables_to_save.items():
            if table_name in raw_tables_in_base: # save tables available in table
                for recid in sorted(table.keys()):
                    if recid in raw_tables_in_base[table_name]:
                        if raw_tables_in_base[table_name][recid] !=  raw_tables_to_save[table_name][recid]:
                            raw_tables_to_update[table_name][recid] = raw_tables_to_save[table_name][recid]
                    else:
                        raw_tables_to_append[table_name][recid] = raw_tables_to_save[table_name][recid]
        self.raw_update(raw_tables_to_update)
        self.raw_append(raw_tables_to_append)   

class XLSXConnector(Connector):
 
    def __init__(self, path, mod, has_header=True):
        super().__init__(path, mod, has_header)
        self.wb_values_only = None

    def load(self):
        self.wb_values_only = openpyxl.load_workbook(filename=self.path, data_only=True)
        self.wb_with_formulas = openpyxl.load_workbook(filename=self.path)
        for wb_sheet in self.wb_values_only:
            sheet = wb_sheet.title
            self.tables[sheet] = {}
            py_table_name = hyperc.xtj.str_to_py(f'{sheet}') # warning only sheet in
            if py_table_name in self.mod.HCT_OBJECTS:
                raise ValueError(f'Error sheet {sheet} already exist')
            header_map = {}
            header_back_map = {}
            header_name_map = {} # map python name to true name
            if self.has_header:
                is_header = True
            else:
                is_header = False
            ThisTable = hyper_etable.meta_table.TableElementMeta(f'{py_table_name}_Class', (object,), {'__table_name__': py_table_name, '__xl_sheet_name__': sheet})
            ThisTable.__annotations__ = {'__table_name__': str, 'addidx': int}
            ThisTable.__header_back_map__ = header_back_map
            ThisTable.__header_name_map__ = header_name_map
            ThisTable.__user_defined_annotations__ = []
            ThisTable.__default_init__ = {}
            ThisTable.__touched_annotations__ = OrderedSet()
            ThisTable.__annotations_type_set__ = collections.defaultdict(set)
            ThisTable.__connector__ = self
            self.mod.__dict__[f'{py_table_name}_Class'] = ThisTable
            self.classes[py_table_name] = ThisTable
            self.classes[py_table_name].__qualname__ = f"{self.mod.__name__}.{py_table_name}_Class"
            self.mod.HCT_OBJECTS[py_table_name] = []
            self.HCT_OBJECTS[py_table_name] = self.mod.HCT_OBJECTS[py_table_name]
            self.objects[py_table_name]={}
            ThisTable.__recid_max__ = 0
            for row in wb_sheet.iter_rows():
                recid = list(row)[0].row
                if ThisTable.__recid_max__ < recid:
                   ThisTable.__recid_max__ = recid
                rec_obj = ThisTable()
                rec_obj.addidx = -1
                if self.has_header:
                    rec_obj.__header_back_map__ = header_back_map
                    rec_obj.__header_name_map__ = header_name_map
                rec_obj.__recid__ = recid
                rec_obj.__table_name__ += f'[{self.path}]{sheet}_{recid}'
                rec_obj.__touched_annotations__ = OrderedSet()
                self.objects[py_table_name][recid] = rec_obj
                self.mod.HCT_OBJECTS[py_table_name].append(rec_obj)
                sheet_name = hyperc.xtj.str_to_py(f"{sheet}") + f'_{recid}'
                if not hasattr(self.mod.DATA, sheet_name):
                    setattr(self.mod.DATA, sheet_name, self.objects[py_table_name][recid])
                    self.mod.StaticObject.__annotations__[sheet_name] = self.classes[py_table_name]
                
                rec_obj.__py_sheet_name__ = sheet_name

                for _cell in row:
                    xl_orig_calculated_value = getattr(_cell, "value", None)

                    letter = _cell.column_letter
                    if is_header:
                        assert xl_orig_calculated_value is not None, "first row can't have empty cell'"
                        header_map[letter] = hyperc.xtj.str_to_py(xl_orig_calculated_value)
                        header_back_map[hyperc.xtj.str_to_py(xl_orig_calculated_value)] = letter
                        header_name_map[hyperc.xtj.str_to_py(xl_orig_calculated_value)] = xl_orig_calculated_value
                        continue
                    elif xl_orig_calculated_value is None:
                        continue
                    if self.has_header:
                        column_name = header_map.get(letter, None)
                        #Skip column with empty header bug #176
                        if column_name is None or column_name == "":
                            continue
                    else:
                        column_name = letter
                    if self.has_header:
                        self.objects[py_table_name][recid].__header_back_map__ = header_back_map

                    self.objects[py_table_name][recid].__touched_annotations__.add(column_name)

                    if xl_orig_calculated_value in ['#NAME?', '#VALUE!']:
                        raise Exception(f"We don't support table with error cell ")
                    if (type(xl_orig_calculated_value) == bool or type(xl_orig_calculated_value) == int or type(xl_orig_calculated_value) == str):
                        setattr(self.objects[py_table_name][recid], column_name, xl_orig_calculated_value)
                        setattr(self.objects[py_table_name][recid], column_name, xl_orig_calculated_value)
                        self.objects[py_table_name][recid].__class__.__annotations__[column_name] = str
                        self.objects[py_table_name][recid].__touched_annotations__.add(column_name) 
                    else:
                        setattr(self.objects[py_table_name][recid], column_name, '')
                        self.objects[py_table_name][recid].__class__.__annotations__[column_name] = str
                        self.objects[py_table_name][recid].__touched_annotations__.add(column_name)
                if is_header:
                    is_header = False
                    continue

                for column_name in header_back_map.keys():
                    if not hasattr(rec_obj, column_name):
                        setattr(self.objects[py_table_name][recid], column_name, '')
                        self.objects[py_table_name][recid].__class__.__annotations__[column_name] = str
                        self.objects[py_table_name][recid].__touched_annotations__.add(column_name)


    def save_all_deprecated(self, out_file=None):
        """Save objects into XLSX file"""
        if out_file is None:
            out_file = self.path
        if self.wb_values_only is None:
            self.wb_with_formulas = openpyxl.Workbook()

        out_dir = pathlib.Path(out_file).parent
        try:
            os.mkdir(out_dir)
        except FileExistsError:
            pass
        for table in self.mod.HCT_OBJECTS.values():
            for row in table:
                sheet_name = row.__xl_sheet_name__
                recid = row.__recid__
                for attr_name in row.__touched_annotations__:
                    if self.has_header:
                        letter = row.__header_back_map__[attr_name]
                    else:
                        letter = attr_name
                    new_value = getattr(row, attr_name)
                    if self.wb_values_only is not None:
                        if sheet_name not in self.wb_values_only:
                            self.wb_values_only.create_sheet(sheet_name)
                            self.wb_with_formulas.create_sheet(sheet_name)
                        elif getattr(self.wb_values_only[sheet_name][f'{letter}{recid}'], "value", None) == new_value:
                            continue
                        self.wb_values_only[sheet_name][f'{letter}{recid}'].value = new_value
                    else:
                        if sheet_name not in self.wb_with_formulas:
                            self.wb_with_formulas.create_sheet(sheet_name)
                        if self.has_header:
                            if getattr(self.wb_with_formulas[sheet_name][f'{letter}1'], "value", None) != row.__header_name_map__[attr_name]:
                                self.wb_with_formulas[sheet_name][f'{letter}1'].value = row.__header_name_map__[attr_name]

                    self.wb_with_formulas[sheet_name][f'{letter}{recid}'].value = new_value
        self.wb_with_formulas.save(out_file)

    def raw_update(self, tables, out_file=None, force_create=True):
        if out_file is None:
            out_file = self.path
        if self.wb_values_only is None:
            self.wb_with_formulas = openpyxl.Workbook()

        for sheet_name, table in tables.items():
            for recid, row in table.items():
                for letter, new_value in row.items():
                    self.wb_with_formulas[sheet_name][f'{letter}{recid}'].value = new_value
        self.wb_with_formulas.save(out_file)

    def raw_append(self, tables, out_file=None):
        self.raw_update(tables, out_file)

    def __str__(self):
        return f'XLSX_FILE_{hyperc.xtj.str_to_py(self.path)}'

class CSVConnector(Connector):

    def __str__(self):
        return f'CSV_FILE_{hyperc.xtj.str_to_py(self.path)}'

class GSheetConnector(XLSXConnector):

    def get_credential(self):
        SCOPES = ['https://www.googleapis.com/auth/drive','https://www.googleapis.com/auth/drive.metadata','https://www.googleapis.com/auth/spreadsheets']

        creds = None
        # The file token.json stores the user's access and refresh tokens, and is
        # created automatically when the authorization flow completes for the first
        # time.
        if os.path.exists('./token.json'):
            creds = Credentials.from_authorized_user_file('token.json', SCOPES)
        # If there are no (valid) credentials available, let the user log in.
        if not creds or not creds.valid:
            bad_token = True
            try:
                if creds and creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                    bad_token = False
            except google.auth.exceptions.RefreshError:
                bad_token = True
            if bad_token:
                flow = InstalledAppFlow.from_client_secrets_file(
                    'credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
            # Save the credentials for the next run
            with open('token.json', 'w') as token:
                token.write(creds.to_json())

    def load(self):
        file_id = self.path
        drive_service = googleapiclient.discovery.build('drive', 'v3', credentials=self.get_credential())

        # request = drive_service.files().get_media(fileId=file_id)
        request = drive_service.files().export_media(fileId=file_id, mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # request = drive_service.files().export_media(fileId=file_id, mimeType='text/csv')

        fh = io.BytesIO()

        downloader = googleapiclient.http.MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            print("Download %d%%." % int(status.progress() * 100))

        self.path = fh
        super().load()

    def save(self, out_file=None):

    # Batch update example
    # {
    #   "range": "Sheet1!B1:D2",
    #   "majorDimension": "ROWS",
    #   "values": [
    #     ["Cost", "Stocked", "Ship Date"],
    #     ["$20.50", "4", "3/1/2016"]
    #   ]
    # }


        batch_update = []

        if out_file is None:
            out_file = self.path
        letter_counter = hyperc.util.letter_index_next()
        for table in self.mod.HCT_OBJECTS.values():
            if self.classes[table].__connector__ is not self:
                continue
            row_values = []
            for row in table:
                sheet_name = row.__xl_sheet_name__
                recid = row.__recid__
                letter_counter=''
                for attr_name in row.__touched_annotations__:
                    letter_counter = hyperc.util.letter_index_next(letter_counter)
                    if self.has_header:
                        letter = row.__header_back_map__[attr_name]
                    else:
                        letter = attr_name

                    while letter_counter != letter:
                        row_values.append(None)

                    new_value = getattr(row, attr_name)
                    # TODO seems useless code (check it)
                    # if sheet_name not in self.wb_values_only:
                    #     self.wb_values_only.create_sheet(sheet_name)
                    #     self.wb_with_formulas.create_sheet(sheet_name)
                    if getattr(self.wb_values_only[sheet_name][f'{letter}{recid}'], "value", None) == new_value:
                        continue
                    row_values.append(new_value)
                    self.wb_values_only[sheet_name][f'{letter}{recid}'].value = new_value
                    self.wb_with_formulas[sheet_name][f'{letter}{recid}'].value = new_value
                batch_update.append({
                    "range": f"{self.classes[table].__xl_sheet_name__}!A1:{letter_counter}{recid}",
                    "majorDimension": "ROWS",
                    "values": row_values
                    })

        sheet_service = googleapiclient.discovery.build('sheets', 'v4', credentials=self.get_credential())
        # batch update
        # https://developers.google.com/sheets/api/reference/rest/v4/spreadsheets.values/batchUpdate

        # Call the Sheets API
        sheet = sheet_service.spreadsheets()
        result = sheet.values().batchUpdate(spreadsheetId=self.path,
                                    range=SAMPLE_RANGE_NAME).execute()


    def __str__(self):
        return f'GSHEET_{hyperc.xtj.str_to_py(self.path)}'

class MSAPIConnector(Connector):
    def __init__(self, path, has_header=True):
        super().__init__(self, path, has_header)
        self.ms_table = hyper_etable.ms_api.get_excel(path)
        # TODO fix code before
        # Load used cell
        for sheet in self.ms_table:
            wb_sheet=self.ms_table[sheet]
            filename = self.path
            # py_table_name = hyperc.xtj.str_to_py(f'[{filename}]{sheet}')
            py_table_name = hyperc.xtj.str_to_py(f'{sheet}') # warning only sheet in 
            header_map = {}
            header_back_map = {}
            if self.has_header:
                is_header = True
            else:
                is_header = False
            ThisTable = hyper_etable.meta_table.TableElementMeta(f'{py_table_name}_Class', (object,), {'__table_name__': py_table_name, '__xl_sheet_name__': sheet})
            ThisTable.__annotations__ = {'__table_name__': str, 'addidx': int}
            ThisTable.__header_back_map__ = header_back_map
            ThisTable.__user_defined_annotations__ = []
            ThisTable.__default_init__ = {}
            ThisTable.__touched_annotations__ = OrderedSet()
            ThisTable.__annotations_type_set__ = collections.defaultdict(set)
            ThisTable.__connector__ = self
            self.mod.__dict__[f'{py_table_name}_Class'] = ThisTable
            self.classes[py_table_name] = ThisTable
            self.classes[py_table_name].__qualname__ = f"{self.session_name}.{py_table_name}_Class"
            self.mod.HCT_OBJECTS[py_table_name] = []
            self.objects[py_table_name]={}
            ThisTable.__recid_max__ = 0
            for row in wb_sheet.iter_rows():
                recid = list(row)[0].row
                if ThisTable.__recid_max__ < recid:
                   ThisTable.__recid_max__ = recid
                rec_obj = ThisTable()
                rec_obj.addidx = -1
                if self.has_header:
                    rec_obj.__header_back_map__ = header_back_map
                rec_obj.__recid__ = recid
                rec_obj.__table_name__ += f'[{filename}]{sheet}_{recid}'
                rec_obj.__touched_annotations__ = OrderedSet()
                self.objects[py_table_name][recid] = rec_obj
                self.mod.HCT_OBJECTS[py_table_name].append(rec_obj)
                sheet_name = hyperc.xtj.str_to_py(f"{sheet}") + f'_{recid}'
                if not hasattr(self.mod.DATA, sheet_name):
                    setattr(self.mod.DATA, sheet_name, self.objects[py_table_name][recid])
                    self.mod.StaticObject.__annotations__[sheet_name] = self.classes[py_table_name]
                
                rec_obj.__py_sheet_name__ = sheet_name

                for _cell in row:
                    xl_orig_calculated_value = getattr(_cell, "value", None)
                    if xl_orig_calculated_value is None:
                        continue
                    letter = _cell.column_letter
                    if is_header:
                        # if xl_orig_calculated_value is None:
                        #     continue
                        header_map[letter] = hyperc.xtj.str_to_py(xl_orig_calculated_value)
                        header_back_map[hyperc.xtj.str_to_py(xl_orig_calculated_value)] = letter
                        continue
                    cell = hyper_etable.cell_resolver.PlainCell(filename=filename, sheet=sheet, letter=letter, number=recid)
                    if self.has_header:
                        column_name = header_map.get(letter, None)
                        #Skip column with empty header bug #176
                        if column_name is None or column_name == "":
                            continue
                    else:
                        column_name = letter
                    if self.has_header:
                        self.objects[py_table_name][recid].__header_back_map__ = header_back_map

                    self.objects[py_table_name][recid].__touched_annotations__.add(column_name)

                    if xl_orig_calculated_value in ['#NAME?', '#VALUE!']:
                        raise Exception(f"We don't support table with error cell {cell}")
                    if (type(xl_orig_calculated_value) == bool or type(xl_orig_calculated_value) == int or type(xl_orig_calculated_value) == str):
                        setattr(self.objects[py_table_name][recid], column_name, xl_orig_calculated_value)
                        setattr(self.objects[py_table_name][recid], column_name, xl_orig_calculated_value)
                        self.objects[py_table_name][recid].__class__.__annotations__[column_name] = str
                        self.objects[py_table_name][recid].__touched_annotations__.add(column_name) 
                    else:
                        setattr(self.objects[py_table_name][recid], column_name, '')
                        self.objects[py_table_name][recid].__class__.__annotations__[column_name] = str
                        self.objects[py_table_name][recid].__touched_annotations__.add(column_name)
                if is_header:
                    is_header = False
                    continue

                for column_name in header_back_map.keys():
                    if not hasattr(rec_obj, column_name):
                        setattr(self.objects[py_table_name][recid], column_name, '')
                        self.objects[py_table_name][recid].__class__.__annotations__[column_name] = str
                        self.objects[py_table_name][recid].__touched_annotations__.add(column_name)

    def __str__(self):
        return f'MSAPI_{hyperc.xtj.str_to_py(self.path)}'

class SQLAlchemyConnector(Connector):

    def load_raw(self):
        raw_tables = {}
        path, tables = self.path
        import sqlalchemy
        engine = sqlalchemy.create_engine(path)
        with engine.connect() as connection:
            for table in tables:
                raw_tables[table] = {}
                result = connection.execute(sqlalchemy.text(f'select * from "{table}"'))
                for row in result:
                    recid = row[0]
                    assert type(recid) is int, "First column should be integer for reqid"
                    raw_tables[table][recid]={}
                    for column_num, column in enumerate(row):
                        if column_num == 0:
                            continue
                        if type(row[column_num]) == bytes :
                            raw_tables[table][recid][row._fields[column_num]] = row[column_num].decode("utf-8") 
                        else:
                            raw_tables[table][recid][row._fields[column_num]] = row[column_num]
        return raw_tables

    def raw_update(self, tables):
        path, _ = self.path
        import sqlalchemy
        engine = sqlalchemy.create_engine(path)

        with engine.connect() as connection:
            inspector = sqlalchemy.inspect(connection)
            inspector.get_table_names() #returns "dow"
            for table in tables:
                if len(tables[table]) == 0 :
                    continue
                columns = inspector.get_columns(table)
                recid_column_name = list(columns)[0]['name']
                for recid, row in tables[table].items():
                    set_column = ", ".join([f'{col} = "{val}"' for col, val in row.items()])
                    query = f'UPDATE {table} SET {set_column} WHERE {recid_column_name} = "{recid}"'
                    connection.execute((query))

    def raw_append(self, tables):
        path, _ = self.path
        import sqlalchemy
        engine = sqlalchemy.create_engine(path)

        with engine.connect() as connection:
            inspector = sqlalchemy.inspect(connection)
            inspector.get_table_names() #returns "dow"
            for table in tables:
                if len(tables[table]) == 0 :
                    continue
                columns = inspector.get_columns(table)
                recid_column_name = list(columns)[0]['name']
                for recid, row in tables[table].items():
                    val = ", ".join([f'"{val}"' for _, val in row.items()])
                    col_name = ", ".join([f'{col}' for col, _ in row.items()])
                    query = f"INSERT INTO {table}({recid_column_name}, {col_name}) VALUES ({recid}, {val})"
                    connection.execute((query))

class MySQLConnector(Connector):
    def load_raw(self):
        raw_tables = {}
        import mysql.connector
        user, password, host, database, tables = self.path
        cnx = mysql.connector.connect(user=user, password=password, host=host, database=database)
        for table in tables:
            cursor = cnx.cursor()

            query = f"SELECT * FROM {table} "

            cursor.execute(query)
            
            raw_tables[table] = {}
            for row in cursor:
                recid = row[0]
                assert type(recid) is int, "First column should be integer for reqid"
                raw_tables[table][recid]={}
                for column_num, column in enumerate(cursor.column_names):
                    if column_num == 0:
                        continue
                    if type(row[column_num]) == bytes :
                        raw_tables[table][recid][column] = row[column_num].decode("utf-8") 
                    else:
                        raw_tables[table][recid][column] = row[column_num]

            cursor.close()
        cnx.close()
        return raw_tables

    def raw_update(self, tables):
        import mysql.connector
        user, password, host, database, _ = self.path
        cnx = mysql.connector.connect(user=user, password=password, host=host, database=database)
        cursor = cnx.cursor()
        for table in tables:
            if len(tables[table]) == 0 :
                continue
            query = (f"DESCRIBE {table} ")
            cursor.execute(query)
            recid_column_name = list(cursor)[0][0]
            query = []
            for recid, row in tables[table].items():
                set_column = ", ".join([f'{col} = "{val}"' for col, val in row.items()])
                query = f'UPDATE {table} SET {set_column} WHERE {recid_column_name} = "{recid}"'
                cursor.execute((query))
        cnx.commit()
        cursor.close()
        cnx.close()


    def raw_append(self, tables):
        import mysql.connector
        user, password, host, database, _ = self.path
        cnx = mysql.connector.connect(user=user, password=password, host=host, database=database)
        cursor = cnx.cursor()
        for table in tables:
            if len(tables[table]) == 0 :
                continue
            query = (f"DESCRIBE {table} ")
            cursor.execute(query)
            recid_column_name = list(cursor)[0][0]
            query = []
            for recid, row in tables[table].items():
                val = ", ".join([f'"{val}"' for col, val in row.items()])
                col_name = ", ".join([f'{col}' for col, val in row.items()])
                query = f"INSERT INTO {table}({recid_column_name}, {col_name}) VALUES ({recid}, {val})"
                cursor.execute(query)
        cnx.commit()
        cursor.close()
        cnx.close()


class AirtableConnector(Connector):
    def load(self):
        BASE_ID, API_KEY, TABLE = self.path

        # response = requests.get(f'https://api.airtable.com/v0/meta/bases/{BASE_ID}/tables?api_key={API_KEY}',).json()
        response = requests.get(f'https://api.airtable.com/v0/{BASE_ID}/{TABLE}?api_key={API_KEY}').json()
        assert 'error' not in response, f"Airtable error {response}"
        # py_table_name = hyperc.xtj.str_to_py(f'[{filename}]{sheet}')
        py_table_name = hyperc.xtj.str_to_py(f'{TABLE}') # warning only sheet in 
        if py_table_name in self.mod.HCT_OBJECTS:
            raise ValueError(f'Error sheet {TABLE} already exist')
        ThisTable = hyper_etable.meta_table.TableElementMeta(f'{py_table_name}_Class', (object,), {'__table_name__': py_table_name, '__xl_sheet_name__': TABLE})
        ThisTable.__annotations__ = {'__table_name__': str, 'addidx': int}
        ThisTable.__user_defined_annotations__ = []
        ThisTable.__default_init__ = {}
        ThisTable.__touched_annotations__ = OrderedSet()
        ThisTable.__annotations_type_set__ = collections.defaultdict(set)
        ThisTable.__connector__ = self
        self.mod.__dict__[f'{py_table_name}_Class'] = ThisTable
        self.classes[py_table_name] = ThisTable
        self.classes[py_table_name].__qualname__ = f"{self.mod.__name__}.{py_table_name}_Class"
        self.mod.HCT_OBJECTS[py_table_name] = []
        self.objects[py_table_name]={}
        ThisTable.__recid_max__ = 0
        recid = 0
        for row in response['records']:
            recid += 1
            if ThisTable.__recid_max__ < recid:
                ThisTable.__recid_max__ = recid
            rec_obj = ThisTable()
            rec_obj.addidx = -1
            rec_obj.__recid__ = recid
            rec_obj.__table_name__ += f'[{BASE_ID}]{TABLE}_{recid}'
            rec_obj.__touched_annotations__ = OrderedSet()
            rec_obj.__xl_sheet_name__ = TABLE
            self.objects[py_table_name][recid] = rec_obj
            self.mod.HCT_OBJECTS[py_table_name].append(rec_obj)
            sheet_name = hyperc.xtj.str_to_py(f"{TABLE}") + f'_{recid}'
            if not hasattr(self.mod.DATA, sheet_name):
                setattr(self.mod.DATA, sheet_name, self.objects[py_table_name][recid])
                self.mod.StaticObject.__annotations__[sheet_name] = self.classes[py_table_name]
            
            rec_obj.__py_sheet_name__ = sheet_name

            for column_name, value in row['fields'].items():

                self.objects[py_table_name][recid].__touched_annotations__.add(column_name)
                if (type(value) == bool or type(value) == int or type(value) == str):
                    setattr(self.objects[py_table_name][recid], column_name, value)
                    setattr(self.objects[py_table_name][recid], column_name, value)
                    self.objects[py_table_name][recid].__class__.__annotations__[column_name] = str
                    self.objects[py_table_name][recid].__touched_annotations__.add(column_name) 
                else:
                    setattr(self.objects[py_table_name][recid], column_name, '')
                    self.objects[py_table_name][recid].__class__.__annotations__[column_name] = str
                    self.objects[py_table_name][recid].__touched_annotations__.add(column_name)
